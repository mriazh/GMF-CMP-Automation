"""Excel report generation with embedded dashboard screenshot."""

import logging
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Config
from .exceptions import ExcelReportError

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates final Excel report with embedded dashboard screenshot."""

    # Configuration for image placement
    IMAGE_TOP_ROWS = 6  # Number of empty rows reserved for the image area
    IMAGE_SPAN_COLS = 15  # Number of columns image should span

    def __init__(self, config: Config):
        self.config = config

    def generate_report(
        self,
        xlsx_path: Path,
        screenshot_path: Path,
        output_path: Path | None = None,
    ) -> Path:
        """Generate final report with dashboard screenshot embedded."""
        logger.info("Generating Excel report: %s -> %s", xlsx_path, output_path or "same file")

        if not xlsx_path.exists():
            raise ExcelReportError(f"Source XLSX not found: {xlsx_path}")
        if not screenshot_path.exists():
            raise ExcelReportError(f"Screenshot not found: {screenshot_path}")

        # Load workbook
        try:
            wb = load_workbook(xlsx_path)
        except zipfile.BadZipFile as e:
            raise ExcelReportError(f"Source XLSX is not a valid XLSX file: {xlsx_path}") from e
        ws = wb.active

        if ws is None:
            raise ExcelReportError("No active worksheet found")

        already_processed = xlsx_path.stem.endswith(" - Edited") and bool(
            getattr(ws, "_images", [])
        )
        if not already_processed:
            self._insert_top_rows(ws, self.IMAGE_TOP_ROWS)
            self._insert_image(ws, screenshot_path)
        else:
            logger.debug("Workbook already contains a dashboard image; preserving layout")

        # Determine output path
        if output_path is None:
            stem = xlsx_path.stem
            if stem.endswith(" - Edited"):
                output_path = xlsx_path
            else:
                output_path = xlsx_path.parent / f"{stem} - Edited.xlsx"

        # Save
        wb.save(output_path)
        logger.info("Report saved to: %s", output_path)

        # Verify
        self._verify_workbook(output_path)

        return output_path

    def _insert_top_rows(self, ws: Worksheet, num_rows: int) -> None:
        """Insert empty rows at the top of the worksheet."""
        logger.debug("Inserting %d rows at top", num_rows)
        ws.insert_rows(1, num_rows)

    def _insert_image(self, ws: Worksheet, image_path: Path) -> None:
        """Insert and size the dashboard screenshot at A1."""
        logger.debug("Inserting image at A1")
        img = XLImage(str(image_path))

        target_width = sum(
            (ws.column_dimensions[get_column_letter(col)].width or 8.43) * 7
            for col in range(1, self.IMAGE_SPAN_COLS + 1)
        )
        target_height = sum(
            (ws.row_dimensions[row].height or 15) * 96 / 72
            for row in range(1, self.IMAGE_TOP_ROWS + 1)
        )

        # Get original image dimensions
        orig_width = img.width
        orig_height = img.height

        # Calculate scale factor to fit within target area while maintaining aspect ratio
        width_scale = target_width / orig_width
        height_scale = target_height / orig_height
        scale = min(width_scale, height_scale) * 0.95  # Slight margin

        img.width = int(orig_width * scale)
        img.height = int(orig_height * scale)

        # Anchor at A1
        ws.add_image(img, "A1")
        logger.debug("Image sized to %dx%d (scale: %.2f)", img.width, img.height, scale)

    def _verify_workbook(self, path: Path) -> None:
        """Verify the workbook can be reopened successfully."""
        logger.debug("Verifying workbook: %s", path)
        try:
            wb = load_workbook(path)
            ws = wb.active
            if ws is None:
                raise ExcelReportError("Verified workbook has no active worksheet")
            images = getattr(ws, "_images", [])
            if len(images) != 1:
                raise ExcelReportError("Verified workbook must contain exactly one dashboard image")
            anchor = images[0].anchor
            if (
                getattr(anchor, "_from", None) is None
                or anchor._from.col != 0
                or anchor._from.row != 0
            ):
                raise ExcelReportError("Dashboard image is not anchored at A1")
            if not images[0].width or not images[0].height:
                raise ExcelReportError("Dashboard image has invalid dimensions")
            if ws.cell(row=7, column=1).value is None:
                raise ExcelReportError("Exported data header is not at row 7")
            logger.debug("Workbook verification successful")
        except Exception as e:
            raise ExcelReportError("Workbook verification failed", str(e)) from e
