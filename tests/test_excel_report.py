"""Tests for Excel report generation."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from cmp_automation.config import Config
from cmp_automation.excel_report import ExcelReportGenerator
from cmp_automation.exceptions import ExcelReportError


class TestExcelReportGenerator:
    """Tests for Excel report generation."""

    @pytest.fixture
    def config(self):
        """Create a test config."""
        return Config(
            cmp_username="test",
            cmp_password="test",
            gmf_email="test@test.com",
            gmf_password="test",
            firefox_profile_dir="/tmp/profile",
            download_dir="/tmp/downloads",
            timezone="Asia/Jakarta",
        )

    @pytest.fixture
    def generator(self, config):
        """Create an ExcelReportGenerator."""
        return ExcelReportGenerator(config)

    @pytest.fixture
    def sample_xlsx(self):
        """Create a sample XLSX file for testing."""
        import os
        import tempfile
        # Create a temp file and close it immediately
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"

            # Add some sample data
            headers = ["ID", "Name", "Status", "Billing Status", "Created"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row in range(2, 11):
                ws.cell(row=row, column=1, value=row - 1)
                ws.cell(row=row, column=2, value=f"Product {row - 1}")
                ws.cell(row=row, column=3, value="Active")
                ws.cell(row=row, column=4, value="Paid")
                ws.cell(row=row, column=5, value="2024-01-15")

            wb.save(path)
            yield Path(path)
        finally:
            try:
                Path(path).unlink(missing_ok=True)
            except Exception:
                pass

    @pytest.fixture
    def sample_screenshot(self):
        """Create a sample screenshot (100x100 pixel PNG)."""
        import os
        import tempfile

        from PIL import Image
        # Create a temp file and close it immediately
        fd, path = tempfile.mkstemp(suffix=".png")
        os.close(fd)
        try:
            img = Image.new('RGB', (100, 100), color='red')
            img.save(path, 'PNG')
            yield Path(path)
        finally:
            try:
                Path(path).unlink(missing_ok=True)
            except Exception:
                pass

    def test_generate_report_creates_output(self, generator, sample_xlsx, sample_screenshot):
        """Test that report generation creates output file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report_with_dashboard.xlsx"
            result = generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            assert result == output_path
            assert result.exists()

    def test_generate_report_auto_naming(self, generator, sample_xlsx, sample_screenshot):
        """Test automatic output naming with ' - Edited' suffix."""
        result = generator.generate_report(sample_xlsx, sample_screenshot)

        assert result.name.endswith(" - Edited.xlsx")
        assert result.exists()

    def test_generate_report_inserts_rows(self, generator, sample_xlsx, sample_screenshot):
        """Test that exactly 6 rows are inserted at top."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            # Reload and check
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            # Original header moves to row 7 (6 inserted rows + 1 header);
            # the original first data row moves to row 8.
            assert ws.cell(row=7, column=1).value == "ID"
            assert ws.cell(row=8, column=1).value == 1

    def test_generate_report_image_at_a1(self, generator, sample_xlsx, sample_screenshot):
        """Test that image is inserted at A1."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            assert len(ws._images) == 1
            img = ws._images[0]
            assert img.anchor._from.row == 0  # Row 1 (0-indexed)
            assert img.anchor._from.col == 0  # Col A (0-indexed)

    def test_generate_report_rows_no_custom_height(self, generator, sample_xlsx, sample_screenshot):
        """Test that rows 1-6 do not receive explicit custom heights."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            for row_num in range(1, 7):
                assert ws.row_dimensions[row_num].height is None, (
                    f"Row {row_num} should not have an explicit custom height"
                )

    def test_generate_report_image_sizing(self, generator, sample_xlsx, sample_screenshot):
        """Test that image is sized to span ~15 cols x 6 rows."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            img = ws._images[0]
            # Image should be scaled (original is 1x1, should be scaled up)
            assert img.width > 1
            assert img.height > 1

    def test_generate_report_image_spans_a_to_o(
        self, generator, sample_xlsx, sample_screenshot
    ):
        """Test that the image is embedded and sized to fit the A:O area."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(output_path)
            ws = wb.active

            assert len(ws._images) == 1
            img = ws._images[0]
            assert img.anchor._from.col == 0
            assert img.anchor._from.row == 0

            target_width = sum(
                (ws.column_dimensions[get_column_letter(col)].width or 8.43) * 7
                for col in range(1, 16)
            )
            target_height = sum(
                (ws.row_dimensions[row].height or 15) * 96 / 72
                for row in range(1, 7)
            )
            # The image is scaled to fit the A:O (15 cols x 6 rows) area.
            assert img.width > 0
            assert img.height > 0
            assert img.width <= target_width
            assert img.height <= target_height

    def test_generate_report_column_widths(self, generator, sample_xlsx, sample_screenshot):
        """Test that existing column widths are preserved."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(output_path)
            ws = wb.active

            for col_idx in range(1, 16):
                col_letter = get_column_letter(col_idx)
                assert ws.column_dimensions[col_letter].width == 13.0

    def test_generate_report_verifies_workbook(self, generator, sample_xlsx, sample_screenshot):
        """Test that generated workbook can be reopened."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            generator.generate_report(sample_xlsx, sample_screenshot, output_path)

            # This should not raise
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            assert wb.active is not None

    def test_missing_source_raises(self, generator, sample_screenshot):
        """Test that missing source XLSX raises error."""
        with pytest.raises(ExcelReportError, match="not found"):
            generator.generate_report(Path("/nonexistent.xlsx"), sample_screenshot)

    def test_missing_screenshot_raises(self, generator, sample_xlsx):
        """Test that missing screenshot raises error."""
        with pytest.raises(ExcelReportError, match="not found"):
            generator.generate_report(sample_xlsx, Path("/nonexistent.png"))

    def test_empty_source_raises(self, generator, sample_screenshot):
        """Test that empty/corrupt source raises error."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"not an xlsx file")
            bad_path = Path(f.name)

        try:
            with pytest.raises(ExcelReportError):
                generator.generate_report(bad_path, sample_screenshot)
        finally:
            bad_path.unlink(missing_ok=True)
